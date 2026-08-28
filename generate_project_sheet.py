#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate professional project management Excel for Quadrotor Project:
"Bám mục tiêu tự động và tránh vật cản trong thời gian thực cho Quadrotors"

Team: Duy (Signal/Localization), Tuân (Control/Guidance), Việt Anh (Embedded/Hardware)
Start: 24/08/2026, Duration: 9 weeks
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import os

# ============================================================================
# CONSTANTS & CONFIG
# ============================================================================
PROJECT_START = datetime(2026, 8, 24)
NUM_WEEKS = 9
PROJECT_END = PROJECT_START + timedelta(weeks=NUM_WEEKS) - timedelta(days=1)

TEAM = {
    "Duy": "Xử lý tín hiệu & Định vị",
    "Tuân": "Điều khiển & Dẫn đường",
    "Việt Anh": "Lập trình nhúng & Triển khai",
}

# Color palette - Professional dark theme inspired
C_DARK_BG = "1B2A4A"
C_HEADER_BG = "0D47A1"
C_HEADER_BG2 = "1565C0"
C_ACCENT_BLUE = "2196F3"
C_ACCENT_TEAL = "00897B"
C_ACCENT_ORANGE = "FF6F00"
C_ACCENT_GREEN = "2E7D32"
C_ACCENT_RED = "C62828"
C_ACCENT_PURPLE = "6A1B9A"
C_ACCENT_AMBER = "FF8F00"
C_LIGHT_BLUE = "BBDEFB"
C_LIGHT_GREEN = "C8E6C9"
C_LIGHT_RED = "FFCDD2"
C_LIGHT_YELLOW = "FFF9C4"
C_LIGHT_ORANGE = "FFE0B2"
C_LIGHT_PURPLE = "E1BEE7"
C_WHITE = "FFFFFF"
C_NEAR_WHITE = "F5F7FA"
C_LIGHT_GRAY = "ECEFF1"
C_MID_GRAY = "B0BEC5"
C_DARK_GRAY = "37474F"
C_BLACK = "212121"

# Fills
FILL_HEADER = PatternFill("solid", fgColor=C_HEADER_BG)
FILL_HEADER2 = PatternFill("solid", fgColor=C_HEADER_BG2)
FILL_SUBHEADER = PatternFill("solid", fgColor=C_ACCENT_TEAL)
FILL_WHITE = PatternFill("solid", fgColor=C_WHITE)
FILL_NEAR_WHITE = PatternFill("solid", fgColor=C_NEAR_WHITE)
FILL_ALT_ROW = PatternFill("solid", fgColor=C_LIGHT_GRAY)
FILL_LIGHT_BLUE = PatternFill("solid", fgColor=C_LIGHT_BLUE)
FILL_LIGHT_GREEN = PatternFill("solid", fgColor=C_LIGHT_GREEN)
FILL_LIGHT_RED = PatternFill("solid", fgColor=C_LIGHT_RED)
FILL_LIGHT_YELLOW = PatternFill("solid", fgColor=C_LIGHT_YELLOW)
FILL_LIGHT_ORANGE = PatternFill("solid", fgColor=C_LIGHT_ORANGE)
FILL_LIGHT_PURPLE = PatternFill("solid", fgColor=C_LIGHT_PURPLE)
FILL_DARK = PatternFill("solid", fgColor=C_DARK_BG)
FILL_ACCENT_GREEN = PatternFill("solid", fgColor=C_ACCENT_GREEN)
FILL_ACCENT_RED = PatternFill("solid", fgColor=C_ACCENT_RED)
FILL_ACCENT_TEAL = PatternFill("solid", fgColor=C_ACCENT_TEAL)
FILL_ACCENT_ORANGE = PatternFill("solid", fgColor=C_ACCENT_ORANGE)
FILL_ACCENT_AMBER = PatternFill("solid", fgColor=C_ACCENT_AMBER)
FILL_ACCENT_PURPLE = PatternFill("solid", fgColor=C_ACCENT_PURPLE)
FILL_DARK_GRAY = PatternFill("solid", fgColor=C_DARK_GRAY)

# Fonts
FONT_TITLE = Font(name="Segoe UI", size=18, bold=True, color=C_WHITE)
FONT_SUBTITLE = Font(name="Segoe UI", size=12, bold=False, color=C_LIGHT_BLUE)
FONT_HEADER = Font(name="Segoe UI Semibold", size=11, bold=True, color=C_WHITE)
FONT_HEADER_SM = Font(name="Segoe UI Semibold", size=10, bold=True, color=C_WHITE)
FONT_SUBHEADER = Font(name="Segoe UI Semibold", size=10, bold=True, color=C_WHITE)
FONT_NORMAL = Font(name="Segoe UI", size=10, color=C_BLACK)
FONT_NORMAL_BOLD = Font(name="Segoe UI", size=10, bold=True, color=C_BLACK)
FONT_SMALL = Font(name="Segoe UI", size=9, color=C_DARK_GRAY)
FONT_LABEL = Font(name="Segoe UI", size=10, bold=True, color=C_DARK_BG)
FONT_VALUE = Font(name="Segoe UI", size=12, bold=True, color=C_HEADER_BG)
FONT_VALUE_BIG = Font(name="Segoe UI", size=20, bold=True, color=C_HEADER_BG)
FONT_LINK = Font(name="Segoe UI", size=10, color=C_ACCENT_BLUE, underline="single")
FONT_WHITE_BOLD = Font(name="Segoe UI", size=11, bold=True, color=C_WHITE)
FONT_RED = Font(name="Segoe UI", size=10, bold=True, color=C_ACCENT_RED)
FONT_GREEN = Font(name="Segoe UI", size=10, bold=True, color=C_ACCENT_GREEN)

# Borders
THIN_BORDER = Border(
    left=Side(style="thin", color=C_MID_GRAY),
    right=Side(style="thin", color=C_MID_GRAY),
    top=Side(style="thin", color=C_MID_GRAY),
    bottom=Side(style="thin", color=C_MID_GRAY),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color=C_HEADER_BG))

# Alignments
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
ALIGN_TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_cell(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    """Apply multiple styles to a cell."""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def write_row(ws, row, data, font=FONT_NORMAL, fill=None, alignment=ALIGN_LEFT, border=THIN_BORDER, start_col=1):
    """Write a row of data with consistent styling."""
    for i, val in enumerate(data):
        cell = ws.cell(row=row, column=start_col + i, value=val)
        style_cell(cell, font=font, fill=fill, alignment=alignment, border=border)


def style_range(ws, min_row, max_row, min_col, max_col, **kwargs):
    """Apply style to a range of cells."""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            style_cell(ws.cell(row=r, column=c), **kwargs)


def week_dates(week_num):
    """Return (start_date, end_date) for a given week number (1-indexed)."""
    start = PROJECT_START + timedelta(weeks=week_num - 1)
    end = start + timedelta(days=6)
    return start, end


# ============================================================================
# TASK DATA — Detailed breakdown
# ============================================================================
def get_tasks():
    """Return comprehensive task list aligned with README.md & user constraints."""
    tasks = [
        # GIAI ĐOẠN 1: Setup & Simulation — Tuần 1-2
        # --- Week 1 ---
        ("GĐ1-T01", "GĐ1: Setup & Sim", "W1", "M1", 
         "Xây dựng mô hình động học 6DOF UAV trên Simulink",
         "Mô hình input: lực/moment, output: pose (x,y,z,φ,θ,ψ). Bao gồm mô hình cánh quạt, trọng lực, khí động.",
         "Tuân", "Duy", "🔴 Critical", "—",
         PROJECT_START, PROJECT_START + timedelta(days=3),
         16, "⬜ Not Started", 0,
         "File .slx mô hình UAV",
         "Mô hình chạy ổn định, output đúng physical unit"),

        ("GĐ1-T02", "GĐ1: Setup & Sim", "W1", "M8",
         "Lập trình thuật toán Standard APF baseline",
         "Triển khai Attractive Potential, Repulsive Potential, xử lý GNRON, lực xoay chống Local Minima.",
         "Tuân", "Duy", "🔴 Critical", "GĐ1-T01",
         PROJECT_START + timedelta(days=2), PROJECT_START + timedelta(days=6),
         20, "⬜ Not Started", 0,
         "Module APF (.m/.slx)",
         "APF hội tụ với ≥ 5 cấu hình vật cản trong mô phỏng"),

        ("GĐ1-T03", "GĐ1: Setup & Sim", "W1", "M2",
         "Lập trình ArUco Detection & PnP Pose Estimation (Python)",
         "Phát hiện ArUco marker (cv2.aruco), tính pose 6DOF (solvePnP) và xuất 2D Bounding Box (u,v,w,h).",
         "Duy", "Tuân", "🔴 Critical", "—",
         PROJECT_START, PROJECT_START + timedelta(days=4),
         16, "⬜ Not Started", 0,
         "Python script ArUco detect",
         "Nhận diện ArUco ổn định trên dataset ảnh thô"),

        ("GĐ1-T04", "GĐ1: Setup & Sim", "W1", "M1",
         "Setup Ubuntu 22.04 + ROS 2 Humble + PX4 SITL + Gazebo + RViz",
         "Thiết lập môi trường mô phỏng ROS 2 Humble, PX4 SITL, Gazebo Garden và RViz visualization.",
         "Duy", "Việt Anh", "🔴 Critical", "—",
         PROJECT_START, PROJECT_START + timedelta(days=3),
         12, "⬜ Not Started", 0,
         "Môi trường mô phỏng sẵn sàng",
         "PX4 SITL + Gazebo + RViz chạy mượt, ROS 2 topic nhận đủ"),

        ("GĐ1-T05", "GĐ1: Setup & Sim", "W1", "M9",
         "[Việt Anh] Đọc & nghiên cứu cấu trúc PX4 Firmware Low-Level Controller",
         "Phân tích module mc_pos_control & mc_att_control của PX4, bộ điều khiển P/PID vị trí & vận tốc.",
         "Việt Anh", "Tuân", "🔴 Critical", "—",
         PROJECT_START + timedelta(days=1), PROJECT_START + timedelta(days=5),
         16, "⬜ Not Started", 0,
         "Báo cáo kiến trúc PX4 mc_pos_control",
         "Hiểu rõ sơ đồ P/PID vị trí/vận tốc & các tham số MPC_*"),

        # --- Week 2 ---
        ("GĐ1-T06", "GĐ1: Setup & Sim", "W2", "M8",
         "Quét tham số Standard APF & Nâng cấp VO-APF (Simulink)",
         "Grid search bộ tham số k_att, k_rep. Nâng cấp lực đẩy phụ thuộc vận tốc tương đối (VO-APF).",
         "Tuân", "Duy", "🟡 High", "GĐ1-T02",
         PROJECT_START + timedelta(days=7), PROJECT_START + timedelta(days=10),
         14, "⬜ Not Started", 0,
         "Bảng tham số APF/VO-APF",
         "Tracking error < 0.3m trên Simulink, quỹ đạo mượt"),

        ("GĐ1-T07", "GĐ1: Setup & Sim", "W2", "M3",
         "Lập trình thuật toán H-Pad Depth Masking (Python/OpenCV)",
         "Sử dụng 2D Bounding Box ArUco dilate 15% để zero-out vùng H-Pad trên ma trận depth.",
         "Duy", "Việt Anh", "🔴 Critical", "GĐ1-T03",
         PROJECT_START + timedelta(days=7), PROJECT_START + timedelta(days=11),
         16, "⬜ Not Started", 0,
         "Module Depth Masking",
         "Zero-out chính xác 100% vùng H-Pad BBox trên depth map"),

        ("GĐ1-T08", "GĐ1: Setup & Sim", "W2", "M2",
         "[Hardware Arrived] Bench-test RealSense D435i + Pi 5 OS",
         "Cài đặt Raspberry Pi OS 64-bit, ROS 2 Humble, realsense2_camera driver. Test luồng RGB+Depth.",
         "Việt Anh", "Duy", "🔴 Critical", "—",
         PROJECT_START + timedelta(days=7), PROJECT_START + timedelta(days=11),
         14, "⬜ Not Started", 0,
         "Pi 5 OS & D435i driver OK",
         "realsense2_camera publish RGB + Depth 30 FPS trên Pi 5"),

        ("GĐ1-T09", "GĐ1: Setup & Sim", "W2", "M10",
         "[Việt Anh] Lập trình ROS 2 Servo Gimbal Control (GPIO PWM)",
         "Viết ROS 2 node xuất xung PWM qua chân GPIO Pi 5 để điều khiển Servo Gimbal góc pitch.",
         "Việt Anh", "Tuân", "🟡 High", "GĐ1-T08",
         PROJECT_START + timedelta(days=9), PROJECT_START + timedelta(days=12),
         12, "⬜ Not Started", 0,
         "ROS 2 gimbal control node",
         "Servo quay mượt 0° -> -90° qua lệnh ROS 2 topic"),

        ("GĐ1-T10", "GĐ1: Setup & Sim", "W2", "M5",
         "[Duy] Thiết kế & Mô phỏng IBVS Landing Controller (Vision Side)",
         "Lập trình thuật toán Image-Based Visual Servoing tính sai lệch pixel tâm marker, đóng gói tín hiệu gửi cho Tuân.",
         "Duy", "Tuân", "🔴 Critical", "GĐ1-T03",
         PROJECT_START + timedelta(days=9), PROJECT_START + timedelta(days=13),
         16, "⬜ Not Started", 0,
         "Module IBVS Landing (Python)",
         "Đóng gói IBVS visual servoing signal gửi cho Tuân"),

        # GIAI ĐOẠN 2: SITL & Bench Test — Tuần 3-4
        # --- Week 3 ---
        ("GĐ2-T01", "GĐ2: SITL & Bench", "W3", "M4/M5",
         "Thiết kế & Lập trình EKF 6-state ước lượng H-Pad State",
         "EKF 6-state (x,y,z,vx,vy,vz) lọc mượt vị trí và ước lượng vận tốc bãi đáp di động H-Pad.",
         "Duy", "Tuân", "🔴 Critical", "GĐ1-T03",
         PROJECT_START + timedelta(days=14), PROJECT_START + timedelta(days=18),
         18, "⬜ Not Started", 0,
         "ROS 2 EKF Node (Python/C++)",
         "Ước lượng với sai số vận tốc < 0.3 m/s"),

        ("GĐ2-T02", "GĐ2: SITL & Bench", "W3", "M8/M9",
         "Port VO-APF Planner & IBVS Landing Controller sang ROS 2",
         "Chuyển đổi thuật toán VO-APF và IBVS Landing từ mô phỏng sang ROS 2 C++/Python packages.",
         "Tuân", "Duy", "🔴 Critical", "GĐ1-T06, GĐ1-T10",
         PROJECT_START + timedelta(days=14), PROJECT_START + timedelta(days=19),
         20, "⬜ Not Started", 0,
         "ROS 2 Planner packages",
         "Chạy mượt trong ROS 2, latency < 20 ms/cycle"),

        ("GĐ2-T03", "GĐ2: SITL & Bench", "W3", "M10",
         "Xây dựng ROS 2 Finite State Machine (FSM 4 States)",
         "Quản lý chuyển vùng 4 trạng thái: SEARCH -> FOLLOW -> APPROACH -> LAND, điều khiển Gimbal tilt.",
         "Việt Anh", "Tuân", "🔴 Critical", "GĐ1-T09",
         PROJECT_START + timedelta(days=14), PROJECT_START + timedelta(days=18),
         14, "⬜ Not Started", 0,
         "ROS 2 FSM node + Diagram",
         "Chuyển đổi 4 trạng thái đúng 100%"),

        ("GĐ2-T04", "GĐ2: SITL & Bench", "W3", "M9",
         "[Việt Anh] Phân tích kịch bản tùy biến PX4 Low-Level Position Controller",
         "Nghiên cứu phương án chỉnh sửa tham số PID hoặc can thiệp code mc_pos_control nếu landing offboard bị lệch.",
         "Việt Anh", "Tuân", "🟡 High", "GĐ1-T05",
         PROJECT_START + timedelta(days=16), PROJECT_START + timedelta(days=20),
         12, "⬜ Not Started", 0,
         "Tài liệu hướng dẫn override PX4 controller",
         "Nắm rõ phương pháp sửa code mc_pos_control nếu land lệch"),

        ("GĐ2-T05", "GĐ2: SITL & Bench", "W3", "M2/M3",
         "Bench-test Vision + Depth Masking + IBVS + EKF trên Pi 5",
         "Chạy toàn bộ Vision pipeline thực tế trên Pi 5 + D435i. Đo FPS, latency và độ chính xác depth mask.",
         "Duy", "Việt Anh", "🔴 Critical", "GĐ1-T07, GĐ1-T08, GĐ1-T10",
         PROJECT_START + timedelta(days=16), PROJECT_START + timedelta(days=20),
         14, "⬜ Not Started", 0,
         "Bench-test report",
         "Vision pipeline đạt ≥ 30 FPS, latency < 100 ms trên Pi 5"),

        # --- Week 4 ---
        ("GĐ2-T06", "GĐ2: SITL & Bench", "W4", "M1-M10",
         "★ MILESTONE 1: Full Gazebo SITL Demo (SEARCH->FOLLOW->LAND)",
         "Chạy mô phỏng khép kín full pipeline trên Gazebo SITL. Quay video demo kết quả.",
         "Cả nhóm", "—", "🔴 Critical", "GĐ2-T02, GĐ2-T03",
         PROJECT_START + timedelta(days=21), PROJECT_START + timedelta(days=23),
         12, "⬜ Not Started", 0,
         "SITL Demo Video + Report",
         "Target follow ≥ 60s, né ≥ 3 vật cản, land error < 0.3m"),

        ("GĐ2-T07", "GĐ2: SITL & Bench", "W4", "M1",
         "[Hardware Arrived] Lắp ráp cơ khí S500 V2 + Wiring UBEC Dual Power",
         "Cả nhóm cùng lắp ráp frame S500 V2, mount Pixhawk 6C, Pi 5, Camera D435i, Gimbal, đi nguồn UBEC kép.",
         "Cả nhóm", "—", "🔴 Critical", "—",
         PROJECT_START + timedelta(days=21), PROJECT_START + timedelta(days=25),
         16, "⬜ Not Started", 0,
         "Drone hardware completed",
         "Cân bằng trọng tâm, 2 nguồn 5.2V FC & 5V/5A Pi 5 riêng biệt"),

        ("GĐ2-T08", "GĐ2: SITL & Bench", "W4", "M9",
         "Calibrate Pixhawk 6C + Bench-test UART micro-XRCE-DDS",
         "Calibrate IMU, Mag, Baro, ESC. Kết nối UART micro-XRCE-DDS Pi 5 <-> Pixhawk 6C baud 921600.",
         "Việt Anh", "Duy", "🔴 Critical", "GĐ2-T07",
         PROJECT_START + timedelta(days=24), PROJECT_START + timedelta(days=27),
         14, "⬜ Not Started", 0,
         "Calibration & Telemetry log",
         "Sensor calib 100% OK, telemetry đệm micro-XRCE-DDS ≥ 30 min"),

        # GIAI ĐOẠN 3: Integration & Flight Test — Tuần 5-6
        # --- Week 5 ---
        ("GĐ3-T01", "GĐ3: Integration & Flight", "W5", "M1-M10",
         "Đóng gói ROS 2 Package & Bench Test Dry Run (No Props)",
         "Test chạy khô không gắn cánh quạt: kiểm tra phản ứng motor khi H-Pad di chuyển, test RC Kill Switch.",
         "Cả nhóm", "—", "🔴 Critical", "GĐ2-T08",
         PROJECT_START + timedelta(days=28), PROJECT_START + timedelta(days=31),
         14, "⬜ Not Started", 0,
         "Dry run video + motor log",
         "Motor phản ứng đúng hướng 100%, Kill switch ngắt < 50 ms"),

        ("GĐ3-T02", "GĐ3: Integration & Flight", "W5", "M1-M10",
         "Test Plan Verification cho từng Module trước khi bay",
         "Nghiệm thu toàn bộ 5 bài Test Plan module an toàn trước khi thực hiện outdoor flight test.",
         "Cả nhóm", "—", "🔴 Critical", "GĐ3-T01",
         PROJECT_START + timedelta(days=31), PROJECT_START + timedelta(days=34),
         12, "⬜ Not Started", 0,
         "Module Test Plan Checklist",
         "Tất cả 5 module vượt qua bài kiểm tra an toàn"),

        # --- Week 6 ---
        ("GĐ3-T03", "GĐ3: Integration & Flight", "W6", "M9",
         "Flight Test #1-2: Pha SEARCH & FOLLOW bãi trống",
         "Bay thử ngoài trời bãi trống. Quadrotor cất cánh, quay tìm ArUco và duy trì FOLLOW ở d_follow = 2-3m.",
         "Cả nhóm", "—", "🔴 Critical", "GĐ3-T02",
         PROJECT_START + timedelta(days=35), PROJECT_START + timedelta(days=38),
         12, "⬜ Not Started", 0,
         "Flight #1-2 log + Video",
         "★ MS2: Bám mục tiêu ≥ 30s, tracking error < 1.0m RMSE"),

        ("GĐ3-T04", "GĐ3: Integration & Flight", "W6", "M9",
         "Fine-tune tham số PX4 Position Controller (MPC_*) nếu cần",
         "Điều chỉnh MPC_XY_P, MPC_Z_P trên Pixhawk 6C nếu đáp ứng bám vị trí thực tế chưa mượt.",
         "Việt Anh", "Tuân", "🟡 High", "GĐ3-T03",
         PROJECT_START + timedelta(days=38), PROJECT_START + timedelta(days=41),
         10, "⬜ Not Started", 0,
         "Updated PX4 param config",
         "Đáp ứng phản hồi vị trí mượt, không bị dội / dao động"),

        # GIAI ĐOẠN 4: Landing & Report — Tuần 7-9
        # --- Week 7 ---
        ("GĐ4-T01", "GĐ4: Landing & Report", "W7", "M8/M9",
         "Flight Test #3-5: Pha FOLLOW + Né vật cản (APF/VO-APF)",
         "Bố trí vật cản mềm trên đường bay. Đánh giá khả năng né vật cản thực tế của VO-APF.",
         "Cả nhóm", "—", "🔴 Critical", "GĐ3-T03",
         PROJECT_START + timedelta(days=42), PROJECT_START + timedelta(days=45),
         14, "⬜ Not Started", 0,
         "Flight #3-5 log + Video",
         "★ MS3: Né vật cản thành công, min clearance > 0.8m"),

        ("GĐ4-T02", "GĐ4: Landing & Report", "W7", "M10",
         "Test pha APPROACH & Gimbal multi-phase tilt angle",
         "Test pha tiếp cận H-Pad, kiểm tra Gimbal tilt mượt xuống -60° và bám tâm marker.",
         "Việt Anh", "Tuân", "🟡 High", "GĐ4-T01",
         PROJECT_START + timedelta(days=45), PROJECT_START + timedelta(days=48),
         10, "⬜ Not Started", 0,
         "Flight APPROACH video",
         "Gimbal tilt mượt xuống -60° khi tiếp cận H-Pad"),

        # --- Week 8 ---
        ("GĐ4-T03", "GĐ4: Landing & Report", "W8", "M1-M10",
         "Flight Test #6-10: Full Pipeline & Precision LAND trên Target Động",
         "Thực hiện trọn vẹn 4 pha SEARCH -> FOLLOW -> APPROACH -> LAND lên H-Pad di chuyển (≤ 5.0 m/s).",
         "Cả nhóm", "—", "🔴 Critical", "GĐ4-T01, GĐ4-T02",
         PROJECT_START + timedelta(days=49), PROJECT_START + timedelta(days=52),
         16, "⬜ Not Started", 0,
         "Flight #6-10 log + Video",
         "★ MS4: Land accuracy < 30cm từ tâm ArUco marker"),

        ("GĐ4-T04", "GĐ4: Landing & Report", "W8", "—",
         "Trích xuất PX4 ulog + ROS 2 bag & Vẽ đồ thị KPI",
         "Phân tích ulog và bag file. Vẽ đồ thị tracking error, obstacle clearance, latency.",
         "Duy", "Tuân", "🔴 Critical", "GĐ4-T03",
         PROJECT_START + timedelta(days=51), PROJECT_START + timedelta(days=54),
         12, "⬜ Not Started", 0,
         "Đồ thị KPI (.png/.pdf)",
         "Đo đủ 7 chỉ số KPI theo yêu cầu đề tài"),

        ("GĐ4-T05", "GĐ4: Landing & Report", "W8", "—",
         "Đóng gói mã nguồn GitHub + Viết tài liệu vận hành",
         "Clean mã nguồn ROS 2, tạo README chi tiết hướng dẫn build, config và deploy.",
         "Việt Anh", "Duy", "🟡 High", "GĐ4-T03",
         PROJECT_START + timedelta(days=51), PROJECT_START + timedelta(days=54),
         10, "⬜ Not Started", 0,
         "Source code & Manual",
         "Code clean, README chi tiết hướng dẫn deploy"),

        # --- Week 9 (Buffer) ---
        ("GĐ4-T06", "GĐ4: Landing & Report", "W9", "—",
         "Hoàn thiện Báo cáo Nghiệm thu + Video Demo chính thức",
         "Viết báo cáo tổng kết đồ án, chỉnh sửa Video Demo 3 phút hoàn chỉnh nghiệm thu.",
         "Cả nhóm", "—", "🔴 Critical", "GĐ4-T04, GĐ4-T05",
         PROJECT_START + timedelta(days=56), PROJECT_START + timedelta(days=60),
         18, "⬜ Not Started", 0,
         "Final Report docx + Video",
         "Báo cáo hoàn chỉnh, Video demo chuyên nghiệp ≥ 3 min"),

        ("GĐ4-T07", "GĐ4: Landing & Report", "W9", "—",
         "Buffer: Bay bổ sung / Dự phòng sự cố",
         "Thời gian dự phòng phát sinh cho các tình huống hư hỏng/cần bay bổ sung data.",
         "Cả nhóm", "—", "🟢 Normal", "—",
         PROJECT_START + timedelta(days=56), PROJECT_START + timedelta(days=62),
         12, "⬜ Not Started", 0,
         "Updated results nếu có",
         "—"),
    ]
    return tasks


# ============================================================================
# RISK DATA
# ============================================================================
def get_risks():
    return [
        ("R01", "Camera depth không chính xác ngoài trời do ánh sáng mạnh/IR interference",
         "High", "High", "Mitigation: Sử dụng stereo matching thay vì structured light depth. Test outdoor trước W5.",
         "Fallback: GPS-only relative positioning", "Duy", "Open", "GĐ1-T06, GĐ3-T05"),

        ("R02", "TensorRT conversion fail hoặc FPS quá thấp trên máy tính nhúng",
         "Medium", "High", "Mitigation: Chuẩn bị model nhẹ hơn (YOLOv8n). Test conversion sớm từ W3.",
         "Fallback: Chạy ONNX Runtime thay TensorRT", "Duy", "Open", "GĐ3-T03"),

        ("R03", "Giao tiếp UART/USB giữa companion computer và PX4 không ổn định",
         "Medium", "Critical", "Mitigation: Test kỹ baudrate, cable quality. Có backup cable.",
         "Fallback: Sử dụng Ethernet/WiFi bridge", "Việt Anh", "Open", "GĐ3-T02"),

        ("R04", "GPS drift gây sai lệch vị trí khi bay outdoor",
         "High", "Medium", "Mitigation: Sử dụng relative position (camera-based) thay vì absolute GPS.",
         "Fallback: Bay trong khu vực có RTK GPS", "Duy", "Open", "GĐ3-T08"),

        ("R05", "Wind disturbance gây mất ổn định khi bay thực",
         "Medium", "High", "Mitigation: Chọn ngày gió nhẹ < 3 m/s. Tune PID cho wind rejection.",
         "Fallback: Bay indoor trong nhà kho/gym lớn", "Tuân", "Open", "GĐ3-T08, GĐ4-T02"),

        ("R06", "Local Minima trong APF - UAV bị kẹt giữa các vật cản",
         "Medium", "Medium", "Mitigation: Modified APF với lực xoay + random perturbation.",
         "Fallback: Switch sang simple waypoint bypass", "Tuân", "Open", "GĐ2-T06"),

        ("R07", "Thời gian 9 tuần không đủ để hoàn thành toàn bộ scope",
         "High", "High", "Mitigation: Prioritize MVP (follow without NN prediction). Weekly review cắt scope sớm.",
         "Fallback: Demo SITL nếu không kịp bay thật", "Cả nhóm", "Open", "GĐ4-T07"),

        ("R08", "Phần cứng hư hỏng khi bay test (crash, ESC cháy...)",
         "Medium", "Critical", "Mitigation: Dry run kỹ. Bay thấp (< 3m) ban đầu. Kill switch luôn sẵn sàng.",
         "Fallback: Có spare parts (ESC, propeller). Budget repair time vào W9.", "Việt Anh", "Open", "GĐ3-T06, GĐ4-T02"),
    ]


# ============================================================================
# SHEET 1: DASHBOARD
# ============================================================================
def create_dashboard(wb):
    if "DASHBOARD" in wb.sheetnames:
        ws = wb["DASHBOARD"]
    else:
        ws = wb.create_sheet("DASHBOARD", 0)
    ws.title = "DASHBOARD"
    ws.sheet_properties.tabColor = C_HEADER_BG

    # Column widths
    col_widths = [3, 22, 20, 20, 20, 20, 20, 3]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Background fill for margin columns
    for r in range(1, 55):
        ws.cell(row=r, column=1).fill = FILL_DARK
        ws.cell(row=r, column=8).fill = FILL_DARK

    # === TITLE BAR ===
    ws.merge_cells("B1:G1")
    ws.merge_cells("B2:G2")
    ws.merge_cells("B3:G3")
    for c in range(2, 8):
        ws.cell(row=1, column=c).fill = FILL_DARK
        ws.cell(row=2, column=c).fill = FILL_DARK
        ws.cell(row=3, column=c).fill = FILL_DARK

    cell = ws.cell(row=2, column=2,
                   value="🚁 QUADROTOR PROJECT — BÁM MỤC TIÊU & TRÁNH VẬT CẢN")
    style_cell(cell, font=FONT_TITLE, fill=FILL_DARK, alignment=ALIGN_CENTER)

    cell = ws.cell(row=3, column=2,
                   value="Automatic Target Following & Obstacle Avoidance in Real-Time")
    style_cell(cell, font=FONT_SUBTITLE, fill=FILL_DARK, alignment=ALIGN_CENTER)

    # === PROJECT INFO CARDS ===
    row = 5
    info_labels = ["Project Start", "Project End", "Duration", "Today", "Days Remaining", "Weeks Remaining"]
    info_values = [
        PROJECT_START.strftime("%d/%m/%Y"),
        PROJECT_END.strftime("%d/%m/%Y"),
        f"{NUM_WEEKS} weeks",
        None,  # formula
        None,  # formula
        None,  # formula
    ]

    # Row 5: Labels
    for c in range(2, 8):
        ws.cell(row=row, column=c).fill = FILL_HEADER
    for i, label in enumerate(info_labels):
        cell = ws.cell(row=row, column=2 + i, value=label)
        style_cell(cell, font=FONT_HEADER_SM, fill=FILL_HEADER, alignment=ALIGN_CENTER, border=THIN_BORDER)

    # Row 6: Values
    row = 6
    for i in range(6):
        cell = ws.cell(row=row, column=2 + i)
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER
        cell.fill = FILL_NEAR_WHITE

    ws.cell(row=6, column=2, value=PROJECT_START.strftime("%d/%m/%Y")).font = FONT_VALUE
    ws.cell(row=6, column=3, value=PROJECT_END.strftime("%d/%m/%Y")).font = FONT_VALUE
    ws.cell(row=6, column=4, value=f"{NUM_WEEKS} tuần").font = FONT_VALUE
    ws.cell(row=6, column=5, value="=TODAY()").font = FONT_VALUE
    ws.cell(row=6, column=5).number_format = "DD/MM/YYYY"
    ws.cell(row=6, column=6).font = FONT_VALUE_BIG
    ws.cell(row=6, column=6, value=f'=MAX(0, DATE(2026,10,25)-TODAY())')
    ws.cell(row=6, column=7).font = FONT_VALUE_BIG
    ws.cell(row=6, column=7, value=f'=ROUNDUP(G6/7, 0)')

    # === TEAM SECTION ===
    row = 8
    ws.merge_cells(f"B{row}:G{row}")
    cell = ws.cell(row=row, column=2, value="👥 TEAM MEMBERS")
    style_cell(cell, font=FONT_HEADER, fill=FILL_HEADER2, alignment=ALIGN_CENTER, border=THIN_BORDER)
    for c in range(3, 8):
        ws.cell(row=row, column=c).fill = FILL_HEADER2
        ws.cell(row=row, column=c).border = THIN_BORDER

    row = 9
    team_headers = ["Member", "Role", "Tasks Assigned", "Tasks Completed", "Completion %"]
    for i, h in enumerate(team_headers):
        cell = ws.cell(row=row, column=2 + i, value=h)
        style_cell(cell, font=FONT_SUBHEADER, fill=FILL_ACCENT_TEAL, alignment=ALIGN_CENTER, border=THIN_BORDER)
    # Leave cols 7 empty but styled
    ws.cell(row=row, column=7).fill = FILL_ACCENT_TEAL
    ws.cell(row=row, column=7).border = THIN_BORDER

    members = [("Duy", "Xử lý tín hiệu & Định vị"), ("Tuân", "Điều khiển & Dẫn đường"), ("Việt Anh", "Lập trình nhúng & Triển khai")]
    for i, (name, role) in enumerate(members):
        r = 10 + i
        fill = FILL_NEAR_WHITE if i % 2 == 0 else FILL_WHITE
        ws.cell(row=r, column=2, value=name)
        style_cell(ws.cell(row=r, column=2), font=FONT_NORMAL_BOLD, fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)
        ws.cell(row=r, column=3, value=role)
        style_cell(ws.cell(row=r, column=3), font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)

        # Formulas counting from TASK_TRACKER
        assigned_formula = f'=COUNTIF(TASK_TRACKER!H:H,"*{name}*")+COUNTIF(TASK_TRACKER!H:H,"Cả nhóm")'
        completed_formula = f'=COUNTIFS(TASK_TRACKER!H:H,"*{name}*",TASK_TRACKER!P:P,"✅ Done")+COUNTIFS(TASK_TRACKER!H:H,"Cả nhóm",TASK_TRACKER!P:P,"✅ Done")'
        pct_formula = f'=IF({get_column_letter(4)}{r}=0,0,{get_column_letter(5)}{r}/{get_column_letter(4)}{r})'

        ws.cell(row=r, column=4, value=assigned_formula)
        style_cell(ws.cell(row=r, column=4), font=FONT_VALUE, fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)
        ws.cell(row=r, column=5, value=completed_formula)
        style_cell(ws.cell(row=r, column=5), font=FONT_GREEN, fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)
        ws.cell(row=r, column=6, value=pct_formula)
        style_cell(ws.cell(row=r, column=6), font=FONT_VALUE, fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)
        ws.cell(row=r, column=6).number_format = '0%'
        ws.cell(row=r, column=7).fill = fill
        ws.cell(row=r, column=7).border = THIN_BORDER

    # === MILESTONE COUNTDOWN ===
    row = 14
    ws.merge_cells(f"B{row}:G{row}")
    cell = ws.cell(row=row, column=2, value="🎯 MILESTONE COUNTDOWN")
    style_cell(cell, font=FONT_HEADER, fill=FILL_HEADER2, alignment=ALIGN_CENTER, border=THIN_BORDER)
    for c in range(3, 8):
        ws.cell(row=row, column=c).fill = FILL_HEADER2
        ws.cell(row=row, column=c).border = THIN_BORDER

    row = 15
    ms_headers = ["Milestone", "Target Date", "Days Left", "KPI", "Status"]
    for i, h in enumerate(ms_headers):
        cell = ws.cell(row=row, column=2 + i, value=h)
        style_cell(cell, font=FONT_SUBHEADER, fill=FILL_ACCENT_TEAL, alignment=ALIGN_CENTER, border=THIN_BORDER)
    ws.cell(row=row, column=7).fill = FILL_ACCENT_TEAL
    ws.cell(row=row, column=7).border = THIN_BORDER

    milestones = [
        ("MS1: Simulink APF hoàn chỉnh", PROJECT_START + timedelta(days=6),
         "Tracking error < 0.3m (sim)", "⬜ Pending"),
        ("MS2: Vision Pipeline 30 FPS", PROJECT_START + timedelta(days=13),
         "mAP > 0.8, 3D error < 0.5m", "⬜ Pending"),
        ("MS3: SITL Follow-Me + Obstacle", PROJECT_START + timedelta(days=27),
         "Bám ≥ 60s, né ≥ 3 vật cản", "⬜ Pending"),
        ("MS4: Hardware Ready", PROJECT_START + timedelta(days=34),
         "Telemetry ổn định ≥ 30 phút", "⬜ Pending"),
        ("MS5: Real Flight Follow", PROJECT_START + timedelta(days=41),
         "Bám ≥ 30s, error < 1.5m", "⬜ Pending"),
        ("MS6: Full System (Obstacle)", PROJECT_START + timedelta(days=48),
         "Min clearance > 0.8m", "⬜ Pending"),
        ("MS7: Final Report", PROJECT_START + timedelta(days=62),
         "Báo cáo + video hoàn chỉnh", "⬜ Pending"),
    ]

    for i, (ms_name, ms_date, kpi, status) in enumerate(milestones):
        r = 16 + i
        fill = FILL_NEAR_WHITE if i % 2 == 0 else FILL_WHITE
        ws.cell(row=r, column=2, value=ms_name)
        style_cell(ws.cell(row=r, column=2), font=FONT_NORMAL_BOLD, fill=fill, alignment=ALIGN_LEFT, border=THIN_BORDER)
        ws.cell(row=r, column=3, value=ms_date)
        style_cell(ws.cell(row=r, column=3), font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)
        ws.cell(row=r, column=3).number_format = "DD/MM/YYYY"

        # Days left formula
        days_formula = f'=MAX(0, C{r}-TODAY())'
        ws.cell(row=r, column=4, value=days_formula)
        style_cell(ws.cell(row=r, column=4), font=FONT_VALUE, fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)

        ws.cell(row=r, column=5, value=kpi)
        style_cell(ws.cell(row=r, column=5), font=FONT_SMALL, fill=fill, alignment=ALIGN_LEFT, border=THIN_BORDER)
        ws.cell(row=r, column=6, value=status)
        style_cell(ws.cell(row=r, column=6), font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)
        ws.cell(row=r, column=7).fill = fill
        ws.cell(row=r, column=7).border = THIN_BORDER

    # Conditional formatting for Days Left - red when < 3
    ws.conditional_formatting.add(
        f"D16:D22",
        CellIsRule(operator="lessThanOrEqual", formula=["3"],
                   fill=PatternFill("solid", fgColor="FFCDD2"),
                   font=Font(bold=True, color=C_ACCENT_RED))
    )
    ws.conditional_formatting.add(
        f"D16:D22",
        CellIsRule(operator="between", formula=["4", "7"],
                   fill=PatternFill("solid", fgColor="FFF9C4"),
                   font=Font(bold=True, color=C_ACCENT_ORANGE))
    )

    # === OVERALL PROGRESS ===
    row = 24
    ws.merge_cells(f"B{row}:G{row}")
    cell = ws.cell(row=row, column=2, value="📊 OVERALL PROGRESS")
    style_cell(cell, font=FONT_HEADER, fill=FILL_HEADER2, alignment=ALIGN_CENTER, border=THIN_BORDER)
    for c in range(3, 8):
        ws.cell(row=row, column=c).fill = FILL_HEADER2
        ws.cell(row=row, column=c).border = THIN_BORDER

    row = 25
    progress_labels = ["Total Tasks", "Completed", "In Progress", "Blocked", "Overall %"]
    for i, label in enumerate(progress_labels):
        cell = ws.cell(row=row, column=2 + i, value=label)
        style_cell(cell, font=FONT_SUBHEADER, fill=FILL_ACCENT_TEAL, alignment=ALIGN_CENTER, border=THIN_BORDER)
    ws.cell(row=row, column=7).fill = FILL_ACCENT_TEAL
    ws.cell(row=row, column=7).border = THIN_BORDER

    row = 26
    formulas = [
        '=COUNTA(TASK_TRACKER!B4:B100)',
        '=COUNTIF(TASK_TRACKER!P4:P100,"✅ Done")',
        '=COUNTIF(TASK_TRACKER!P4:P100,"🔵 In Progress")',
        '=COUNTIF(TASK_TRACKER!P4:P100,"🔴 Blocked")',
        '=IF(B26=0,0,C26/B26)',
    ]
    fonts_list = [FONT_VALUE_BIG, FONT_GREEN, FONT_VALUE, FONT_RED, FONT_VALUE_BIG]
    for i, (formula, fnt) in enumerate(zip(formulas, fonts_list)):
        cell = ws.cell(row=row, column=2 + i, value=formula)
        style_cell(cell, font=fnt, fill=FILL_NEAR_WHITE, alignment=ALIGN_CENTER, border=THIN_BORDER)
    ws.cell(row=26, column=6).number_format = '0%'
    ws.cell(row=row, column=7).fill = FILL_NEAR_WHITE
    ws.cell(row=row, column=7).border = THIN_BORDER

    # === PHASE PROGRESS ===
    row = 28
    ws.merge_cells(f"B{row}:G{row}")
    cell = ws.cell(row=row, column=2, value="📈 PHASE PROGRESS")
    style_cell(cell, font=FONT_HEADER, fill=FILL_HEADER2, alignment=ALIGN_CENTER, border=THIN_BORDER)
    for c in range(3, 8):
        ws.cell(row=row, column=c).fill = FILL_HEADER2
        ws.cell(row=row, column=c).border = THIN_BORDER

    row = 29
    phase_headers = ["Phase", "Total Tasks", "Done", "Progress %", "Weeks"]
    for i, h in enumerate(phase_headers):
        cell = ws.cell(row=row, column=2 + i, value=h)
        style_cell(cell, font=FONT_SUBHEADER, fill=FILL_ACCENT_TEAL, alignment=ALIGN_CENTER, border=THIN_BORDER)
    ws.cell(row=row, column=7).fill = FILL_ACCENT_TEAL
    ws.cell(row=row, column=7).border = THIN_BORDER

    phases = [
        ("GĐ1: Simulation & APF", "GĐ1*", "W1-W2"),
        ("GĐ2: Tracking & Follow", "GĐ2*", "W3-W4"),
        ("GĐ3: Hardware & Bench Test", "GĐ3*", "W5-W6"),
        ("GĐ4: Thực nghiệm & Báo cáo", "GĐ4*", "W7-W9"),
    ]

    for i, (phase_name, pattern, weeks) in enumerate(phases):
        r = 30 + i
        fill = FILL_NEAR_WHITE if i % 2 == 0 else FILL_WHITE
        ws.cell(row=r, column=2, value=phase_name)
        style_cell(ws.cell(row=r, column=2), font=FONT_NORMAL_BOLD, fill=fill, alignment=ALIGN_LEFT, border=THIN_BORDER)

        total_f = f'=COUNTIF(TASK_TRACKER!C4:C100,"{pattern}")'
        done_f = f'=COUNTIFS(TASK_TRACKER!C4:C100,"{pattern}",TASK_TRACKER!P4:P100,"✅ Done")'
        pct_f = f'=IF(C{r}=0,0,D{r}/C{r})'

        ws.cell(row=r, column=3, value=total_f)
        style_cell(ws.cell(row=r, column=3), font=FONT_VALUE, fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)
        ws.cell(row=r, column=4, value=done_f)
        style_cell(ws.cell(row=r, column=4), font=FONT_GREEN, fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)
        ws.cell(row=r, column=5, value=pct_f)
        style_cell(ws.cell(row=r, column=5), font=FONT_VALUE, fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)
        ws.cell(row=r, column=5).number_format = '0%'
        ws.cell(row=r, column=6, value=weeks)
        style_cell(ws.cell(row=r, column=6), font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)
        ws.cell(row=r, column=7).fill = fill
        ws.cell(row=r, column=7).border = THIN_BORDER

    # Data bar for progress %
    ws.conditional_formatting.add(
        f"E30:E33",
        DataBarRule(start_type="num", start_value=0,
                    end_type="num", end_value=1,
                    color=C_ACCENT_GREEN)
    )

    # === ALERTS / NOTES ===
    row = 35
    ws.merge_cells(f"B{row}:G{row}")
    cell = ws.cell(row=row, column=2, value="⚠️ ALERTS & NOTES")
    style_cell(cell, font=FONT_HEADER, fill=FILL_ACCENT_ORANGE, alignment=ALIGN_CENTER, border=THIN_BORDER)
    for c in range(3, 8):
        ws.cell(row=row, column=c).fill = FILL_ACCENT_ORANGE
        ws.cell(row=row, column=c).border = THIN_BORDER

    alerts = [
        "⏰ Overdue tasks:",
        '=COUNTIFS(TASK_TRACKER!P4:P100,"<>✅ Done",TASK_TRACKER!M4:M100,"<",TODAY()) & " task(s) overdue"',
        "",
        "🔴 Blocked tasks:",
        '=COUNTIF(TASK_TRACKER!P4:P100,"🔴 Blocked") & " task(s) blocked"',
        "",
        "📋 Notes: Weekly review mỗi thứ 6. Update status trước khi họp.",
    ]
    for i, alert in enumerate(alerts):
        r = 36 + i
        ws.merge_cells(f"B{r}:G{r}")
        cell = ws.cell(row=r, column=2, value=alert)
        style_cell(cell, font=FONT_NORMAL if "Notes" not in str(alert) else FONT_SMALL,
                   fill=FILL_LIGHT_YELLOW, alignment=ALIGN_LEFT, border=THIN_BORDER)
        for c in range(3, 8):
            ws.cell(row=r, column=c).fill = FILL_LIGHT_YELLOW
            ws.cell(row=r, column=c).border = THIN_BORDER

    # Print settings
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.freeze_panes = "B5"

    return ws


# ============================================================================
# SHEET 2: TASK_TRACKER
# ============================================================================
def create_task_tracker(wb):
    ws = wb.create_sheet("TASK_TRACKER")
    ws.sheet_properties.tabColor = C_ACCENT_TEAL

    # Headers
    headers = [
        "No.", "Task ID", "Giai đoạn", "Week", "Module",
        "Tên nhiệm vụ", "Mô tả chi tiết", "Owner", "Reviewer",
        "Priority", "Predecessor", "Planned Start", "Planned End",
        "Actual Start", "Actual End", "Status", "% Complete",
        "Countdown\n(days)", "Effort\n(hours)", "Deliverable",
        "Acceptance Criteria", "Notes / Blockers"
    ]

    col_widths = [5, 10, 22, 6, 8, 35, 45, 12, 12, 12, 14, 13, 13, 13, 13, 15, 10, 10, 8, 25, 30, 30]

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title rows
    ws.merge_cells("A1:V1")
    cell = ws.cell(row=1, column=1, value="📋 TASK TRACKER — Quadrotor Follow-Me & Obstacle Avoidance Project")
    style_cell(cell, font=FONT_TITLE, fill=FILL_DARK, alignment=ALIGN_CENTER)
    for c in range(2, 23):
        ws.cell(row=1, column=c).fill = FILL_DARK

    ws.merge_cells("A2:V2")
    cell = ws.cell(row=2, column=1,
                   value=f"Start: {PROJECT_START.strftime('%d/%m/%Y')} | End: {PROJECT_END.strftime('%d/%m/%Y')} | Team: Duy · Tuân · Việt Anh")
    style_cell(cell, font=FONT_SUBTITLE, fill=FILL_DARK, alignment=ALIGN_CENTER)
    for c in range(1, 23):
        ws.cell(row=2, column=c).fill = FILL_DARK

    # Header row
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i, value=h)
        style_cell(cell, font=FONT_HEADER_SM, fill=FILL_HEADER, alignment=ALIGN_CENTER, border=THIN_BORDER)

    # Data rows
    tasks = get_tasks()
    for idx, task in enumerate(tasks):
        r = 4 + idx
        (task_id, phase, week, module, name, desc, owner, reviewer, priority,
         predecessor, p_start, p_end, effort, status, pct_complete,
         deliverable, acceptance) = task

        fill = FILL_NEAR_WHITE if idx % 2 == 0 else FILL_WHITE

        # Phase-based left accent
        if "GĐ1" in phase:
            phase_fill = FILL_LIGHT_BLUE
        elif "GĐ2" in phase:
            phase_fill = FILL_LIGHT_GREEN
        elif "GĐ3" in phase:
            phase_fill = FILL_LIGHT_ORANGE
        else:
            phase_fill = FILL_LIGHT_PURPLE

        data = [
            idx + 1, task_id, phase, week, module, name, desc, owner, reviewer,
            priority, predecessor, p_start, p_end,
            None, None,  # Actual Start/End - to be filled
            status, pct_complete / 100 if pct_complete else 0,
        ]

        for i, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=i, value=val)
            current_fill = phase_fill if i == 3 else fill
            font = FONT_NORMAL
            align = ALIGN_LEFT if i in (6, 7) else ALIGN_CENTER

            if i == 6:  # Task name
                font = FONT_NORMAL_BOLD
            elif i == 10:  # Priority
                if "Critical" in str(val):
                    font = Font(name="Segoe UI", size=10, bold=True, color=C_ACCENT_RED)
                elif "High" in str(val):
                    font = Font(name="Segoe UI", size=10, bold=True, color=C_ACCENT_ORANGE)
                else:
                    font = Font(name="Segoe UI", size=10, color=C_ACCENT_GREEN)

            style_cell(cell, font=font, fill=current_fill, alignment=align, border=THIN_BORDER)

        # Date formatting
        for col in [12, 13, 14, 15]:
            ws.cell(row=r, column=col).number_format = "DD/MM/YYYY"

        # Countdown formula: =MAX(0, M{r} - TODAY())
        countdown_f = f'=IF(O{r}="","=MAX(0,M{r}-TODAY())","")'
        # Simpler: just show days until planned end
        ws.cell(row=r, column=18, value=f'=IF(P{r}="✅ Done","—",MAX(0,M{r}-TODAY()))')
        style_cell(ws.cell(row=r, column=18), font=FONT_NORMAL_BOLD, fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)

        # Effort
        ws.cell(row=r, column=19, value=effort)
        style_cell(ws.cell(row=r, column=19), font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)

        # % Complete formatting
        ws.cell(row=r, column=17).number_format = '0%'

        # Deliverable
        ws.cell(row=r, column=20, value=deliverable)
        style_cell(ws.cell(row=r, column=20), font=FONT_SMALL, fill=fill, alignment=ALIGN_TOP_LEFT, border=THIN_BORDER)

        # Acceptance criteria
        ws.cell(row=r, column=21, value=acceptance)
        style_cell(ws.cell(row=r, column=21), font=FONT_SMALL, fill=fill, alignment=ALIGN_TOP_LEFT, border=THIN_BORDER)

        # Notes
        ws.cell(row=r, column=22, value="")
        style_cell(ws.cell(row=r, column=22), font=FONT_SMALL, fill=fill, alignment=ALIGN_TOP_LEFT, border=THIN_BORDER)

    num_tasks = len(tasks)
    last_data_row = 3 + num_tasks

    # Data validations
    # Status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"⬜ Not Started,🔵 In Progress,✅ Done,🔴 Blocked,⏰ Overdue,⏸ On Hold"',
        allow_blank=True
    )
    status_dv.error = "Please select a valid status"
    status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(status_dv)
    status_dv.add(f"P4:P{last_data_row}")

    # Priority dropdown
    priority_dv = DataValidation(
        type="list",
        formula1='"🔴 Critical,🟡 High,🟢 Normal,⚪ Low"',
        allow_blank=True
    )
    ws.add_data_validation(priority_dv)
    priority_dv.add(f"J4:J{last_data_row}")

    # Owner dropdown
    owner_dv = DataValidation(
        type="list",
        formula1='"Duy,Tuân,Việt Anh,Cả nhóm"',
        allow_blank=True
    )
    ws.add_data_validation(owner_dv)
    owner_dv.add(f"H4:H{last_data_row}")

    # Reviewer dropdown
    reviewer_dv = DataValidation(
        type="list",
        formula1='"Duy,Tuân,Việt Anh,—"',
        allow_blank=True
    )
    ws.add_data_validation(reviewer_dv)
    reviewer_dv.add(f"I4:I{last_data_row}")

    # Conditional formatting for Status column
    ws.conditional_formatting.add(
        f"P4:P{last_data_row}",
        CellIsRule(operator="equal", formula=['"✅ Done"'],
                   fill=FILL_LIGHT_GREEN, font=FONT_GREEN)
    )
    ws.conditional_formatting.add(
        f"P4:P{last_data_row}",
        CellIsRule(operator="equal", formula=['"🔴 Blocked"'],
                   fill=FILL_LIGHT_RED, font=FONT_RED)
    )
    ws.conditional_formatting.add(
        f"P4:P{last_data_row}",
        CellIsRule(operator="equal", formula=['"🔵 In Progress"'],
                   fill=FILL_LIGHT_BLUE)
    )
    ws.conditional_formatting.add(
        f"P4:P{last_data_row}",
        CellIsRule(operator="equal", formula=['"⏰ Overdue"'],
                   fill=PatternFill("solid", fgColor="FF8A80"),
                   font=Font(bold=True, color="B71C1C"))
    )

    # Countdown conditional formatting
    ws.conditional_formatting.add(
        f"R4:R{last_data_row}",
        CellIsRule(operator="lessThanOrEqual", formula=["2"],
                   fill=FILL_LIGHT_RED, font=FONT_RED)
    )
    ws.conditional_formatting.add(
        f"R4:R{last_data_row}",
        CellIsRule(operator="between", formula=["3", "5"],
                   fill=FILL_LIGHT_YELLOW, font=Font(bold=True, color=C_ACCENT_AMBER))
    )

    # Data bar for % Complete
    ws.conditional_formatting.add(
        f"Q4:Q{last_data_row}",
        DataBarRule(start_type="num", start_value=0,
                    end_type="num", end_value=1,
                    color=C_ACCENT_GREEN)
    )

    # Freeze panes
    ws.freeze_panes = "G4"

    # Auto filter
    ws.auto_filter.ref = f"A3:V{last_data_row}"

    # Row height
    for r in range(4, last_data_row + 1):
        ws.row_dimensions[r].height = 40

    return ws


# ============================================================================
# SHEET 3: GANTT VIEW
# ============================================================================
def create_gantt_view(wb):
    ws = wb.create_sheet("GANTT_VIEW")
    ws.sheet_properties.tabColor = C_ACCENT_BLUE

    # Column A-C: Task info, then D onwards: days
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 10

    # Title
    total_days = NUM_WEEKS * 7
    last_col = get_column_letter(4 + total_days)

    ws.merge_cells(f"A1:{last_col}1")
    cell = ws.cell(row=1, column=1, value="📅 GANTT CHART — 9-Week Timeline")
    style_cell(cell, font=FONT_TITLE, fill=FILL_DARK, alignment=ALIGN_CENTER)
    for c in range(2, 5 + total_days):
        ws.cell(row=1, column=c).fill = FILL_DARK

    # Week headers (row 2)
    ws.cell(row=2, column=1, value="No.").font = FONT_HEADER_SM
    ws.cell(row=2, column=1).fill = FILL_HEADER
    ws.cell(row=2, column=1).alignment = ALIGN_CENTER
    ws.cell(row=2, column=1).border = THIN_BORDER
    ws.cell(row=2, column=2, value="Task ID").font = FONT_HEADER_SM
    ws.cell(row=2, column=2).fill = FILL_HEADER
    ws.cell(row=2, column=2).alignment = ALIGN_CENTER
    ws.cell(row=2, column=2).border = THIN_BORDER
    ws.cell(row=2, column=3, value="Task Name").font = FONT_HEADER_SM
    ws.cell(row=2, column=3).fill = FILL_HEADER
    ws.cell(row=2, column=3).alignment = ALIGN_CENTER
    ws.cell(row=2, column=3).border = THIN_BORDER
    ws.cell(row=2, column=4, value="Owner").font = FONT_HEADER_SM
    ws.cell(row=2, column=4).fill = FILL_HEADER
    ws.cell(row=2, column=4).alignment = ALIGN_CENTER
    ws.cell(row=2, column=4).border = THIN_BORDER

    # Day headers (row 3) and Week headers (row 2)
    for day_idx in range(total_days):
        col = 5 + day_idx
        date = PROJECT_START + timedelta(days=day_idx)
        week_num = day_idx // 7 + 1

        # Narrow columns for days
        ws.column_dimensions[get_column_letter(col)].width = 3

        # Week header (merge per week)
        if day_idx % 7 == 0:
            start_col = col
            end_col = col + 6
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
            cell = ws.cell(row=2, column=start_col, value=f"W{week_num}")
            week_colors = [C_HEADER_BG, C_HEADER_BG2, C_ACCENT_TEAL, C_ACCENT_PURPLE,
                           C_HEADER_BG, C_HEADER_BG2, C_ACCENT_TEAL, C_ACCENT_PURPLE, C_HEADER_BG]
            style_cell(cell, font=FONT_HEADER_SM,
                       fill=PatternFill("solid", fgColor=week_colors[(week_num - 1) % len(week_colors)]),
                       alignment=ALIGN_CENTER, border=THIN_BORDER)

        # Day number
        cell = ws.cell(row=3, column=col, value=date.day)
        is_weekend = date.weekday() >= 5
        day_fill = PatternFill("solid", fgColor="E0E0E0") if is_weekend else FILL_NEAR_WHITE
        style_cell(cell, font=Font(name="Segoe UI", size=7, color=C_DARK_GRAY),
                   fill=day_fill, alignment=ALIGN_CENTER, border=THIN_BORDER)

    # Row 3 headers for A-D
    for c, label in zip([1, 2, 3, 4], ["", "", "", ""]):
        ws.cell(row=3, column=c).fill = FILL_HEADER2
        ws.cell(row=3, column=c).border = THIN_BORDER

    # Task rows
    tasks = get_tasks()
    phase_fills_gantt = {
        "GĐ1": PatternFill("solid", fgColor="42A5F5"),
        "GĐ2": PatternFill("solid", fgColor="66BB6A"),
        "GĐ3": PatternFill("solid", fgColor="FFA726"),
        "GĐ4": PatternFill("solid", fgColor="AB47BC"),
    }

    for idx, task in enumerate(tasks):
        r = 4 + idx
        task_id = task[0]
        phase = task[1]
        name = task[4]
        owner = task[6]
        p_start = task[10]
        p_end = task[11]

        fill = FILL_NEAR_WHITE if idx % 2 == 0 else FILL_WHITE

        ws.cell(row=r, column=1, value=idx + 1)
        style_cell(ws.cell(row=r, column=1), font=FONT_SMALL, fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)
        ws.cell(row=r, column=2, value=task_id)
        style_cell(ws.cell(row=r, column=2), font=FONT_SMALL, fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)
        ws.cell(row=r, column=3, value=name)
        style_cell(ws.cell(row=r, column=3), font=Font(name="Segoe UI", size=8, color=C_BLACK),
                   fill=fill, alignment=ALIGN_LEFT, border=THIN_BORDER)
        ws.cell(row=r, column=4, value=owner)
        style_cell(ws.cell(row=r, column=4), font=Font(name="Segoe UI", size=8, bold=True),
                   fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)

        # Fill gantt bars
        phase_key = phase[:3]
        bar_fill = phase_fills_gantt.get(phase_key, FILL_LIGHT_BLUE)

        for day_idx in range(total_days):
            col = 5 + day_idx
            date = PROJECT_START + timedelta(days=day_idx)
            is_weekend = date.weekday() >= 5

            cell = ws.cell(row=r, column=col)
            cell.border = Border(
                left=Side(style="hair", color="E0E0E0"),
                right=Side(style="hair", color="E0E0E0"),
                top=Side(style="hair", color="E0E0E0"),
                bottom=Side(style="hair", color="E0E0E0"),
            )

            if p_start <= date <= p_end:
                cell.fill = bar_fill
                # Milestone markers
                if "MILESTONE" in name or "★" in name:
                    if date == p_end:
                        cell.value = "◆"
                        cell.font = Font(name="Segoe UI", size=8, bold=True, color=C_WHITE)
                        cell.alignment = ALIGN_CENTER
            elif is_weekend:
                cell.fill = PatternFill("solid", fgColor="F5F5F5")

        ws.row_dimensions[r].height = 22

    # Legend
    legend_row = 4 + len(tasks) + 2
    ws.merge_cells(f"A{legend_row}:D{legend_row}")
    ws.cell(row=legend_row, column=1, value="Legend:").font = FONT_NORMAL_BOLD

    legends = [
        ("GĐ1: Simulation & APF", "42A5F5"),
        ("GĐ2: Tracking & Follow", "66BB6A"),
        ("GĐ3: Hardware & Bench", "FFA726"),
        ("GĐ4: Thực nghiệm & Báo cáo", "AB47BC"),
        ("◆ = Milestone", "FFFFFF"),
    ]
    for i, (label, color) in enumerate(legends):
        r = legend_row + 1 + i
        ws.cell(row=r, column=2, value="██")
        ws.cell(row=r, column=2).font = Font(color=color)
        ws.cell(row=r, column=3, value=label)
        ws.cell(row=r, column=3).font = FONT_SMALL

    ws.freeze_panes = "E4"
    return ws


# ============================================================================
# SHEET 4: RISK LOG
# ============================================================================
def create_risk_log(wb):
    ws = wb.create_sheet("RISK_LOG")
    ws.sheet_properties.tabColor = C_ACCENT_RED

    headers = [
        "Risk ID", "Mô tả rủi ro", "Probability", "Impact",
        "Risk Score", "Mitigation Plan", "Contingency Plan",
        "Owner", "Status", "Related Tasks", "Last Updated"
    ]
    col_widths = [10, 45, 12, 12, 12, 40, 35, 12, 12, 18, 13]

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells("A1:K1")
    cell = ws.cell(row=1, column=1, value="⚠️ RISK REGISTER — Quadrotor Project")
    style_cell(cell, font=FONT_TITLE, fill=FILL_DARK, alignment=ALIGN_CENTER)
    for c in range(2, 12):
        ws.cell(row=1, column=c).fill = FILL_DARK

    # Subtitle
    ws.merge_cells("A2:K2")
    cell = ws.cell(row=2, column=1, value="Proactive risk management — Required for hardware flight testing")
    style_cell(cell, font=FONT_SUBTITLE, fill=FILL_DARK, alignment=ALIGN_CENTER)
    for c in range(2, 12):
        ws.cell(row=2, column=c).fill = FILL_DARK

    # Headers
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i, value=h)
        style_cell(cell, font=FONT_HEADER_SM, fill=FILL_ACCENT_RED, alignment=ALIGN_CENTER, border=THIN_BORDER)

    # Data
    risks = get_risks()
    for idx, risk in enumerate(risks):
        r = 4 + idx
        risk_id, desc, prob, impact, mitigation, contingency, owner, status, related = risk

        fill = FILL_NEAR_WHITE if idx % 2 == 0 else FILL_WHITE

        # Risk score mapping
        prob_map = {"Low": 1, "Medium": 2, "High": 3}
        impact_map = {"Low": 1, "Medium": 2, "High": 3, "Critical": 4}
        score = prob_map.get(prob, 1) * impact_map.get(impact, 1)

        if score >= 9:
            score_label = f"🔴 {score} (Extreme)"
            score_fill = FILL_LIGHT_RED
        elif score >= 6:
            score_label = f"🟡 {score} (High)"
            score_fill = FILL_LIGHT_YELLOW
        elif score >= 3:
            score_label = f"🟠 {score} (Medium)"
            score_fill = FILL_LIGHT_ORANGE
        else:
            score_label = f"🟢 {score} (Low)"
            score_fill = FILL_LIGHT_GREEN

        data = [risk_id, desc, prob, impact, score_label, mitigation, contingency, owner, status, related, None]
        for i, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=i, value=val)
            current_fill = score_fill if i == 5 else fill
            font = FONT_NORMAL
            align = ALIGN_LEFT if i in (2, 6, 7) else ALIGN_CENTER
            if i == 1:
                font = FONT_NORMAL_BOLD
            style_cell(cell, font=font, fill=current_fill, alignment=align, border=THIN_BORDER)

        # Date format for Last Updated
        ws.cell(row=r, column=11).number_format = "DD/MM/YYYY"

    num_risks = len(risks)
    last_row = 3 + num_risks

    # Data validations
    prob_dv = DataValidation(type="list", formula1='"Low,Medium,High"')
    ws.add_data_validation(prob_dv)
    prob_dv.add(f"C4:C{last_row}")

    impact_dv = DataValidation(type="list", formula1='"Low,Medium,High,Critical"')
    ws.add_data_validation(impact_dv)
    impact_dv.add(f"D4:D{last_row}")

    status_dv = DataValidation(type="list", formula1='"Open,Mitigated,Occurred,Closed"')
    ws.add_data_validation(status_dv)
    status_dv.add(f"I4:I{last_row}")

    owner_dv = DataValidation(type="list", formula1='"Duy,Tuân,Việt Anh,Cả nhóm"')
    ws.add_data_validation(owner_dv)
    owner_dv.add(f"H4:H{last_row}")

    # Conditional formatting for status
    ws.conditional_formatting.add(
        f"I4:I{last_row}",
        CellIsRule(operator="equal", formula=['"Occurred"'],
                   fill=FILL_LIGHT_RED, font=FONT_RED)
    )
    ws.conditional_formatting.add(
        f"I4:I{last_row}",
        CellIsRule(operator="equal", formula=['"Closed"'],
                   fill=FILL_LIGHT_GREEN, font=FONT_GREEN)
    )

    # Risk matrix reference table
    matrix_row = last_row + 3
    ws.merge_cells(f"A{matrix_row}:D{matrix_row}")
    ws.cell(row=matrix_row, column=1, value="📊 Risk Matrix Reference").font = FONT_NORMAL_BOLD

    matrix_headers = ["", "Low Impact (1)", "Medium (2)", "High (3)", "Critical (4)"]
    mr = matrix_row + 1
    for i, h in enumerate(matrix_headers):
        ws.cell(row=mr, column=1 + i, value=h)
        style_cell(ws.cell(row=mr, column=1 + i), font=FONT_HEADER_SM, fill=FILL_HEADER2, alignment=ALIGN_CENTER, border=THIN_BORDER)

    probs = ["High (3)", "Medium (2)", "Low (1)"]
    scores_matrix = [[3, 6, 9, 12], [2, 4, 6, 8], [1, 2, 3, 4]]
    for pi, (p_label, scores) in enumerate(zip(probs, scores_matrix)):
        r = mr + 1 + pi
        ws.cell(row=r, column=1, value=p_label)
        style_cell(ws.cell(row=r, column=1), font=FONT_NORMAL_BOLD, fill=FILL_NEAR_WHITE, alignment=ALIGN_CENTER, border=THIN_BORDER)
        for si, score in enumerate(scores):
            cell = ws.cell(row=r, column=2 + si, value=score)
            if score >= 9:
                sfill = FILL_LIGHT_RED
            elif score >= 6:
                sfill = FILL_LIGHT_YELLOW
            elif score >= 3:
                sfill = FILL_LIGHT_ORANGE
            else:
                sfill = FILL_LIGHT_GREEN
            style_cell(cell, font=FONT_NORMAL_BOLD, fill=sfill, alignment=ALIGN_CENTER, border=THIN_BORDER)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:K{last_row}"

    for r in range(4, last_row + 1):
        ws.row_dimensions[r].height = 45

    return ws


# ============================================================================
# SHEET 5: WEEKLY REPORT
# ============================================================================
def create_weekly_report(wb):
    ws = wb.create_sheet("WEEKLY_REPORT")
    ws.sheet_properties.tabColor = C_ACCENT_PURPLE

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 3
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 3

    # Title
    ws.merge_cells("A1:G1")
    cell = ws.cell(row=1, column=1, value="📝 WEEKLY REPORT TEMPLATE")
    style_cell(cell, font=FONT_TITLE, fill=FILL_DARK, alignment=ALIGN_CENTER)
    for c in range(2, 8):
        ws.cell(row=1, column=c).fill = FILL_DARK

    ws.merge_cells("A2:G2")
    cell = ws.cell(row=2, column=1, value="Fill in before each Friday review meeting")
    style_cell(cell, font=FONT_SUBTITLE, fill=FILL_DARK, alignment=ALIGN_CENTER)
    for c in range(2, 8):
        ws.cell(row=2, column=c).fill = FILL_DARK

    # Create template for each week
    current_row = 4

    for week in range(1, NUM_WEEKS + 1):
        w_start, w_end = week_dates(week)

        # Week header
        ws.merge_cells(f"B{current_row}:F{current_row}")
        header_text = f"📅 WEEK {week}  ({w_start.strftime('%d/%m')} — {w_end.strftime('%d/%m/%Y')})"
        cell = ws.cell(row=current_row, column=2, value=header_text)
        style_cell(cell, font=FONT_WHITE_BOLD, fill=FILL_HEADER, alignment=ALIGN_CENTER, border=THIN_BORDER)
        for c in [3, 4, 5, 6]:
            ws.cell(row=current_row, column=c).fill = FILL_HEADER
            ws.cell(row=current_row, column=c).border = THIN_BORDER
        current_row += 1

        # Review date
        ws.cell(row=current_row, column=2, value="Review Date:")
        style_cell(ws.cell(row=current_row, column=2), font=FONT_LABEL, fill=FILL_NEAR_WHITE, alignment=ALIGN_LEFT, border=THIN_BORDER)
        review_date = w_start + timedelta(days=4)  # Friday
        ws.cell(row=current_row, column=3, value=review_date)
        style_cell(ws.cell(row=current_row, column=3), font=FONT_NORMAL, fill=FILL_NEAR_WHITE, alignment=ALIGN_LEFT, border=THIN_BORDER)
        ws.cell(row=current_row, column=3).number_format = "DD/MM/YYYY"

        ws.cell(row=current_row, column=5, value="Attendees:")
        style_cell(ws.cell(row=current_row, column=5), font=FONT_LABEL, fill=FILL_NEAR_WHITE, alignment=ALIGN_LEFT, border=THIN_BORDER)
        ws.cell(row=current_row, column=6, value="Duy, Tuân, Việt Anh")
        style_cell(ws.cell(row=current_row, column=6), font=FONT_NORMAL, fill=FILL_NEAR_WHITE, alignment=ALIGN_LEFT, border=THIN_BORDER)
        current_row += 1

        # Sections - Left column
        sections_left = [
            ("✅ Completed This Week", FILL_LIGHT_GREEN, 5),
            ("📋 Planned for Next Week", FILL_LIGHT_BLUE, 5),
        ]
        sections_right = [
            ("🚧 Blockers & Issues", FILL_LIGHT_RED, 5),
            ("💡 Decisions Made", FILL_LIGHT_YELLOW, 5),
        ]

        for (left_title, left_fill, left_rows), (right_title, right_fill, right_rows) in zip(sections_left, sections_right):
            # Section headers
            ws.cell(row=current_row, column=2, value=left_title)
            style_cell(ws.cell(row=current_row, column=2), font=FONT_SUBHEADER,
                       fill=FILL_ACCENT_TEAL if "Completed" in left_title else FILL_HEADER2,
                       alignment=ALIGN_LEFT, border=THIN_BORDER)
            ws.cell(row=current_row, column=3).fill = (FILL_ACCENT_TEAL if "Completed" in left_title else FILL_HEADER2)
            ws.cell(row=current_row, column=3).border = THIN_BORDER

            ws.cell(row=current_row, column=5, value=right_title)
            style_cell(ws.cell(row=current_row, column=5), font=FONT_SUBHEADER,
                       fill=FILL_ACCENT_RED if "Blockers" in right_title else FILL_ACCENT_AMBER,
                       alignment=ALIGN_LEFT, border=THIN_BORDER)
            ws.cell(row=current_row, column=6).fill = (FILL_ACCENT_RED if "Blockers" in right_title else FILL_ACCENT_AMBER)
            ws.cell(row=current_row, column=6).border = THIN_BORDER
            current_row += 1

            # Empty rows for filling in
            for line in range(left_rows):
                ws.cell(row=current_row + line, column=2, value=f"{line + 1}.")
                style_cell(ws.cell(row=current_row + line, column=2), font=FONT_SMALL, fill=left_fill, alignment=ALIGN_LEFT, border=THIN_BORDER)
                ws.cell(row=current_row + line, column=3).fill = left_fill
                ws.cell(row=current_row + line, column=3).border = THIN_BORDER

                ws.cell(row=current_row + line, column=5, value=f"{line + 1}.")
                style_cell(ws.cell(row=current_row + line, column=5), font=FONT_SMALL, fill=right_fill, alignment=ALIGN_LEFT, border=THIN_BORDER)
                ws.cell(row=current_row + line, column=6).fill = right_fill
                ws.cell(row=current_row + line, column=6).border = THIN_BORDER

            current_row += left_rows

        # Action Items section (full width)
        ws.merge_cells(f"B{current_row}:F{current_row}")
        ws.cell(row=current_row, column=2, value="🎯 Action Items")
        style_cell(ws.cell(row=current_row, column=2), font=FONT_SUBHEADER, fill=FILL_ACCENT_PURPLE,
                   alignment=ALIGN_LEFT, border=THIN_BORDER)
        for c in range(3, 7):
            ws.cell(row=current_row, column=c).fill = FILL_ACCENT_PURPLE
            ws.cell(row=current_row, column=c).border = THIN_BORDER
        current_row += 1

        # Action items sub-headers
        ai_headers = ["", "Action", "Owner", "", "Deadline"]
        for i, h in enumerate(ai_headers):
            cell = ws.cell(row=current_row, column=2 + i, value=h)
            style_cell(cell, font=FONT_HEADER_SM, fill=FILL_DARK_GRAY, alignment=ALIGN_CENTER, border=THIN_BORDER)
        current_row += 1

        for line in range(3):
            ws.cell(row=current_row, column=2, value=f"{line + 1}.")
            for c in range(2, 7):
                style_cell(ws.cell(row=current_row, column=c), font=FONT_SMALL,
                           fill=FILL_LIGHT_PURPLE, alignment=ALIGN_LEFT, border=THIN_BORDER)
            current_row += 1

        # Spacer
        current_row += 2

    ws.freeze_panes = "A4"
    return ws


# ============================================================================
# MAIN
# ============================================================================
def main():
    target_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "VDT_project.xlsx")
    MANAGED_SHEETS = {"DASHBOARD", "TASK_TRACKER", "GANTT_VIEW", "RISK_LOG", "WEEKLY_REPORT"}

    if os.path.exists(target_path):
        print(f"Loading existing workbook from: {target_path}")
        wb = openpyxl.load_workbook(target_path)
        for s_name in list(wb.sheetnames):
            if s_name in MANAGED_SHEETS:
                del wb[s_name]
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    print("Creating DASHBOARD...")
    create_dashboard(wb)

    print("Creating TASK_TRACKER...")
    create_task_tracker(wb)

    print("Creating GANTT_VIEW...")
    create_gantt_view(wb)

    print("Creating RISK_LOG...")
    create_risk_log(wb)

    print("Creating WEEKLY_REPORT...")
    create_weekly_report(wb)

    # Reorder sheets: MANAGED_SHEETS first, then any preserved sheets (like 'List đồ cần mua')
    ordered_sheets = []
    for s_name in ["DASHBOARD", "TASK_TRACKER", "GANTT_VIEW", "RISK_LOG", "WEEKLY_REPORT"]:
        if s_name in wb.sheetnames:
            ordered_sheets.append(wb[s_name])
    for s_name in wb.sheetnames:
        if s_name not in MANAGED_SHEETS:
            ordered_sheets.append(wb[s_name])
    wb._sheets = ordered_sheets

    try:
        wb.save(target_path)
        print(f"\n[OK] File saved successfully: {target_path}")
    except PermissionError:
        fallback_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "VDT_project_v2.xlsx")
        wb.save(fallback_path)
        print(f"\n[OK] Target file locked. Saved as fallback: {fallback_path}")

    safe_sheetnames = [s.encode('ascii', 'ignore').decode() if not s.isascii() else s for s in wb.sheetnames]
    print(f"Final Sheets: {safe_sheetnames}")
    print(f"Project: {PROJECT_START.strftime('%d/%m/%Y')} - {PROJECT_END.strftime('%d/%m/%Y')} ({NUM_WEEKS} weeks)")
    print(f"Team: Duy, Tuan, Viet Anh")
    print(f"Total tasks: {len(get_tasks())}")
    print(f"Total risks: {len(get_risks())}")


if __name__ == "__main__":
    main()
